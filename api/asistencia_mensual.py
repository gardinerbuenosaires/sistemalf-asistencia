from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from typing import Optional
from datetime import date, timedelta
from collections import defaultdict
from db.database import db_session
from auth.core import require_permiso

router = APIRouter(prefix="/api/asistencia", tags=["asistencia_mensual"])

# ── Mapeo estado evaluador → letra planilla ────────────────────────────────────
_EST_SIMPLE = {
    "ok": "I", "salida_anticipada": "I", "sin_salida": "I", "sin_horario": "I",
    "tarde": "T", "tarde_y_salida_anticipada": "T", "tarde_y_sin_salida": "T",
    "ausente": "A!!!", "franco": "F", "ft": "FT",
    "b1_ausente": "A!!!", "b2_ausente": "A!!!",
    "nf": "NF", "duda": "!",
}
_EST_B1 = {**_EST_SIMPLE, "b1_ausente": "A!!!", "b2_ausente": "I"}
_EST_B2 = {
    "ok": "I", "salida_anticipada": "I", "sin_salida": "I", "sin_horario": "I",
    "tarde": "I",                        # tarde afecta la entrada del b1
    "tarde_y_salida_anticipada": "I",
    "tarde_y_sin_salida": "I",
    "ausente": "A!!!", "franco": "F", "ft": "FT",
    "b1_ausente": "I", "b2_ausente": "A!!!",
    "nf": "NF", "duda": "!",
}

CONTABLES     = {"I", "T", "F", "FT", "FD", "V", "L", "LSG", "S", "@", "NF"}
LETRAS_VALIDAS = {"ILT", "LSG", "L", "E", "V", "S", "FT", "FD", "F", "@", "NF", "A", "CP"}
DIAS_SEMANA   = ["lu", "ma", "mi", "ju", "vi", "sá", "do"]


# ── Resolución de una celda/bloque ────────────────────────────────────────────
def _resolver(eid, fecha, bloque, f_ing, f_egr, nov_map, ali_set, res_map, plan_map):
    if f_ing and fecha < f_ing:
        return {"letra": "O", "tipo": "antes_ingreso"}
    if f_egr and fecha >= f_egr:
        return {"letra": "X", "tipo": "liquidacion"}

    # 1. Novedad manual: bloque específico primero, luego bloque=0 (día entero)
    nov = nov_map.get((eid, fecha, bloque))
    if nov is None and bloque != 0:
        nov = nov_map.get((eid, fecha, 0))
    if nov:
        if nov["tipo"] == "CP":
            r = {"letra": "I", "tipo": "normal", "cp": True, "nov_id": nov["id"]}
        else:
            r = {"letra": nov["tipo"], "tipo": "normal", "nov_id": nov["id"]}
        if nov.get("descripcion"):
            r["descripcion"] = nov["descripcion"]
        if nov.get("creado_por"):
            r["creado_por"] = nov["creado_por"]
        return r

    # 2. Aliviada (solo bloques 1 y 2)
    if bloque in (1, 2) and (eid, fecha, bloque) in ali_set:
        return {"letra": "@", "tipo": "normal"}

    # 3. Resultado automático del evaluador
    res = res_map.get((eid, fecha))
    if res:
        # Bloque MT (no aplica): mostrar dash o highlight si vino inesperadamente
        if bloque == 1 and res.get("b1_mt"):
            if res.get("b1_entrada"):
                return {"letra": "MT", "tipo": "mt_trabajado"}
            return {"letra": "MT", "tipo": "mt"}
        if bloque == 2 and res.get("b2_mt"):
            if res.get("b2_entrada"):
                return {"letra": "MT", "tipo": "mt_trabajado"}
            return {"letra": "MT", "tipo": "mt"}

        estado = res["estado"]

        if estado == "no_iniciado":
            return {"letra": None, "tipo": "no_iniciado"}

        if estado.endswith("_b2_ni"):
            if bloque == 2:
                return {"letra": None, "tipo": "no_iniciado"}
            base = estado[:-6]  # strip "_b2_ni"
            mapa = _EST_B1 if bloque == 1 else _EST_SIMPLE
            letra = mapa.get(base)
            if letra:
                return {"letra": letra, "tipo": "normal"}

        # Duda causada por un bloque con solo salida: el otro bloque completo muestra como presente
        if estado == "duda":
            if bloque == 1 and res.get("b1_entrada"):
                return {"letra": "I", "tipo": "normal"}
            if bloque == 2 and res.get("b2_entrada"):
                return {"letra": "I", "tipo": "normal"}

        # Si b1 es MT, el estado representa solo b2 evaluado como bloque simple → usar _EST_B1 para bloque 2
        if bloque == 2 and res.get("b1_mt"):
            mapa = _EST_B1
        else:
            mapa = _EST_B1 if bloque == 1 else (_EST_B2 if bloque == 2 else _EST_SIMPLE)
        letra  = mapa.get(estado)
        if letra:
            ft_pendiente = estado == "ft" and not res["horario_id"]
            return {"letra": letra, "tipo": "normal", **({"ft_pendiente": True} if ft_pendiente else {})}

    # 4. Planificacion (futuro o sin procesar)
    plan = plan_map.get((eid, fecha))
    if plan:
        if plan["es_franco"]:
            return {"letra": "F", "tipo": "normal"}
        if plan["horario_id"]:
            return {"letra": None, "tipo": "pendiente"}

    return {"letra": None, "tipo": "sin_plan"}


def _fmt(v: float):
    """Devuelve int si es entero, float si tiene decimal."""
    return int(v) if v == int(v) else v


def _calcular_control(fechas, f_egr, celdas, cortado, f0, f1, hoy_str):
    dias_duda          = []
    dias_faltantes     = []
    dias_ausentes_sc   = []
    dias_activos       = 0

    for fecha in fechas:
        celda = celdas[fecha]
        if cortado or celda.get("es_cortado_dia"):
            bloques = [celda["b1"], celda["b2"]]
        else:
            bloques = [celda]

        # Día entero fuera del período activo del empleado → ignorar
        if all(b["tipo"] in ("antes_ingreso", "liquidacion") for b in bloques):
            continue

        dias_activos += 1
        dia_num = int(fecha[8:])

        if any(b.get("letra") == "!" for b in bloques):
            dias_duda.append(dia_num)
            continue  # duda ≠ faltante

        if any(b.get("letra") == "A!!!" for b in bloques):
            dias_ausentes_sc.append(dia_num)
            continue  # ausente sin confirmar ≠ faltante

        if any(b["tipo"] in ("pendiente", "sin_plan", "no_iniciado") for b in bloques):
            dias_faltantes.append(dia_num)

    f1_mas1 = (date.fromisoformat(f1) + timedelta(days=1)).isoformat()
    tiene_egreso_mes = bool(f_egr) and f0 <= f_egr <= f1_mas1

    # Ventana de alerta: últimos 5 días del mes
    ventana_desde = (date.fromisoformat(f1) - timedelta(days=4)).isoformat()
    en_ventana    = hoy_str >= ventana_desde

    extra = {"dias_ausentes_sc": dias_ausentes_sc} if dias_ausentes_sc else {}

    if dias_activos < 3:
        return {"estado": "vacio", **extra}
    if not dias_duda and not dias_faltantes:
        if dias_ausentes_sc:
            return {"estado": "ausentes_sc", **extra}
        return {"estado": "liquidacion" if tiene_egreso_mes else "completado"}
    if dias_duda and not dias_faltantes:
        return {"estado": "revisar", "dias_duda": dias_duda, **extra}
    if not dias_duda:
        # Solo días faltantes: detalle solo en la ventana final
        if en_ventana:
            return {"estado": "faltan", "dias_faltantes": dias_faltantes, **extra}
        return {"estado": "en_curso", **extra}
    # Hay duda Y faltantes: siempre mostrar duda, faltantes solo en ventana
    if en_ventana:
        return {"estado": "mixto", "dias_duda": dias_duda, "dias_faltantes": dias_faltantes}
    return {"estado": "revisar", "dias_duda": dias_duda}


def get_control_estados_batch(conn, eids: list, f0: str, f1: str) -> dict:
    """
    Devuelve {empleado_id: estado_control} para una lista de empleados en el período f0-f1.
    Usa la misma lógica que la planilla (_resolver + _calcular_control).
    Premios lo importa para saber si el empleado tiene asistencia completada.
    """
    if not eids:
        return {}

    fechas = []
    d = date.fromisoformat(f0)
    last = date.fromisoformat(f1)
    while d <= last:
        fechas.append(d.isoformat())
        d += timedelta(days=1)

    ph = ",".join("?" * len(eids))

    emp_rows = conn.execute(
        f"SELECT id, fecha_ingreso, fecha_egreso FROM empleados WHERE id IN ({ph})", eids
    ).fetchall()

    # Tipo de horario predominante por empleado (para saber si es cortado)
    cortado_rows = conn.execute(f"""
        SELECT p.empleado_id, h.tipo
        FROM planificacion p
        JOIN horarios h ON h.id = p.horario_id
        WHERE p.empleado_id IN ({ph}) AND p.fecha >= ? AND p.fecha <= ?
          AND p.horario_id IS NOT NULL
        ORDER BY p.fecha DESC
    """, (*eids, f0, f1)).fetchall()
    cortado_map = {}
    for r in cortado_rows:
        if r["empleado_id"] not in cortado_map:
            cortado_map[r["empleado_id"]] = (r["tipo"] == "cortado")

    planes = conn.execute(
        f"SELECT p.empleado_id, p.fecha, p.es_franco, p.horario_id, h.tipo AS horario_tipo "
        f"FROM planificacion p LEFT JOIN horarios h ON h.id=p.horario_id "
        f"WHERE p.empleado_id IN ({ph}) AND p.fecha>=? AND p.fecha<=?",
        (*eids, f0, f1)
    ).fetchall()

    resultados = conn.execute(
        f"SELECT empleado_id, fecha, estado, horario_id, b1_mt, b2_mt, b1_entrada, b2_entrada "
        f"FROM resultados_dia WHERE empleado_id IN ({ph}) AND fecha>=? AND fecha<=?",
        (*eids, f0, f1)
    ).fetchall()

    novedades = conn.execute(
        f"SELECT id, empleado_id, fecha, bloque, tipo, descripcion, creado_por "
        f"FROM novedades WHERE empleado_id IN ({ph}) AND fecha>=? AND fecha<=?",
        (*eids, f0, f1)
    ).fetchall()

    aliviadas = conn.execute(
        f"SELECT empleado_id, fecha, bloque FROM aliviadas "
        f"WHERE empleado_id IN ({ph}) AND fecha>=? AND fecha<=?",
        (*eids, f0, f1)
    ).fetchall()

    plan_map = {(r["empleado_id"], r["fecha"]): dict(r) for r in planes}
    res_map  = {
        (r["empleado_id"], r["fecha"]): {
            "estado": r["estado"], "horario_id": r["horario_id"],
            "b1_mt": r["b1_mt"], "b2_mt": r["b2_mt"],
            "b1_entrada": r["b1_entrada"], "b2_entrada": r["b2_entrada"],
        }
        for r in resultados
    }
    nov_map = {}
    for n in novedades:
        nov_map[(n["empleado_id"], n["fecha"], n["bloque"])] = {
            "id": n["id"], "tipo": n["tipo"],
            "descripcion": n["descripcion"], "creado_por": n["creado_por"],
        }
    ali_set = {(a["empleado_id"], a["fecha"], a["bloque"]) for a in aliviadas}

    hoy_str = date.today().isoformat()
    resultado = {}

    for emp in emp_rows:
        eid   = emp["id"]
        f_ing = (emp["fecha_ingreso"] or "")[:10]
        f_egr = (emp["fecha_egreso"]  or "")[:10]
        cortado = cortado_map.get(eid, False)

        celdas = {}
        for fecha in fechas:
            plan = plan_map.get((eid, fecha))
            dia_cortado = plan and plan.get("horario_tipo") == "cortado"
            if cortado or dia_cortado:
                b1 = _resolver(eid, fecha, 1, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                b2 = _resolver(eid, fecha, 2, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                celdas[fecha] = {"b1": b1, "b2": b2, "es_cortado_dia": bool(dia_cortado)}
            else:
                c = _resolver(eid, fecha, 0, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                celdas[fecha] = c

        ctrl = _calcular_control(fechas, f_egr, celdas, cortado, f0, f1, hoy_str)
        resultado[eid] = ctrl["estado"]

    return resultado


def _build_dias(fechas, feriados):
    return [
        {
            "fecha": f,
            "num":   int(f[8:]),
            "dow":   DIAS_SEMANA[date.fromisoformat(f).weekday()],
            "feriado": f in feriados,
            "domingo": date.fromisoformat(f).weekday() == 6,
        }
        for f in fechas
    ]


# ── Lógica de datos del mes (compartida entre /mensual y /mensual/excel) ───────
def _asistencia_datos(mes: str) -> dict:
    try:
        año, m = int(mes[:4]), int(mes[5:7])
        primer_dia = date(año, m, 1)
        ultimo_dia = (
            date(año + 1, 1, 1) - timedelta(days=1) if m == 12
            else date(año, m + 1, 1) - timedelta(days=1)
        )
    except Exception:
        raise HTTPException(400, "Formato de mes inválido (YYYY-MM)")

    fechas = []
    d = primer_dia
    while d <= ultimo_dia:
        fechas.append(d.isoformat())
        d += timedelta(days=1)
    f0, f1 = fechas[0], fechas[-1]
    mes_ant = f"{año}-{m-1:02d}" if m > 1 else f"{año-1}-12"

    with db_session() as conn:
        feriados = {
            r["fecha"][:10] for r in conn.execute(
                "SELECT fecha FROM feriados WHERE fecha >= ? AND fecha <= ?", (f0, f1)
            ).fetchall()
        }

        # Empleados: activos + desvinculados cuyo fecha_egreso cae dentro del mes
        emp_rows = conn.execute("""
            WITH ult AS (
                SELECT p.empleado_id, h.tipo, t.nombre AS turno_nombre,
                       ROW_NUMBER() OVER (
                           PARTITION BY p.empleado_id ORDER BY p.fecha DESC
                       ) AS rn
                FROM planificacion p
                JOIN horarios h ON h.id = p.horario_id
                LEFT JOIN turnos t ON t.id = h.turno_id
                WHERE p.fecha >= ? AND p.fecha <= ? AND p.horario_id IS NOT NULL
            )
            SELECT e.id, e.user_id, e.nombre, e.apellido, c.nombre AS cargo,
                   e.fecha_ingreso, e.fecha_egreso, e.en_dispositivo, e.activo,
                   u.tipo AS hipo, u.turno_nombre
            FROM empleados e
            LEFT JOIN cargos c ON c.id = e.cargo_id
            LEFT JOIN ult u ON u.empleado_id = e.id AND u.rn = 1
            WHERE e.tipo != 'acceso'
              AND (e.activo = 1
               OR (e.fecha_egreso > ? AND e.fecha_egreso <= ?))
            ORDER BY e.apellido, e.nombre
        """, (f0, f1, f0, f1)).fetchall()

        if not emp_rows:
            return {"mes": mes, "dias": _build_dias(fechas, feriados),
                    "TM": [], "TN": [], "CO": [], "ST": []}

        eids = [r["id"] for r in emp_rows]
        ph   = ",".join("?" * len(eids))

        planes = conn.execute(
            f"SELECT p.empleado_id, p.fecha, p.es_franco, p.horario_id, h.tipo AS horario_tipo "
            f"FROM planificacion p "
            f"LEFT JOIN horarios h ON h.id = p.horario_id "
            f"WHERE p.empleado_id IN ({ph}) AND p.fecha>=? AND p.fecha<=?",
            (*eids, f0, f1)
        ).fetchall()

        resultados = conn.execute(
            f"SELECT empleado_id, fecha, estado, horario_id, b1_mt, b2_mt, b1_entrada, b2_entrada "
            f"FROM resultados_dia WHERE empleado_id IN ({ph}) AND fecha>=? AND fecha<=?",
            (*eids, f0, f1)
        ).fetchall()

        novedades = conn.execute(
            f"SELECT id, empleado_id, fecha, bloque, tipo, descripcion, creado_por "
            f"FROM novedades WHERE empleado_id IN ({ph}) AND fecha>=? AND fecha<=?",
            (*eids, f0, f1)
        ).fetchall()

        aliviadas = conn.execute(
            f"SELECT empleado_id, fecha, bloque "
            f"FROM aliviadas WHERE empleado_id IN ({ph}) AND fecha>=? AND fecha<=?",
            (*eids, f0, f1)
        ).fetchall()

        saldos = conn.execute(
            f"SELECT empleado_id, saldo FROM saldo_francos "
            f"WHERE empleado_id IN ({ph}) AND mes=?",
            (*eids, mes_ant)
        ).fetchall()

        fichajes_manuales = conn.execute(
            f"SELECT empleado_id, date(timestamp) AS fecha "
            f"FROM fichajes "
            f"WHERE empleado_id IN ({ph}) AND date(timestamp)>=? AND date(timestamp)<=? AND es_manual=1 "
            f"GROUP BY empleado_id, date(timestamp)",
            (*eids, f0, f1)
        ).fetchall()

    # Lookup dicts
    plan_map = {(r["empleado_id"], r["fecha"]): dict(r) for r in planes}
    res_map  = {
        (r["empleado_id"], r["fecha"]): {
            "estado": r["estado"], "horario_id": r["horario_id"],
            "b1_mt": r["b1_mt"], "b2_mt": r["b2_mt"],
            "b1_entrada": r["b1_entrada"], "b2_entrada": r["b2_entrada"],
        }
        for r in resultados
    }

    fm_count: dict[int, int] = defaultdict(int)
    for f in fichajes_manuales:
        fm_count[f["empleado_id"]] += 1
    nov_map  = {}
    comentarios_map: dict[int, list] = defaultdict(list)
    for n in novedades:
        nov_map[(n["empleado_id"], n["fecha"], n["bloque"])] = {
            "id": n["id"], "tipo": n["tipo"], "descripcion": n["descripcion"],
            "creado_por": n["creado_por"],
        }
        comentarios_map[n["empleado_id"]].append({
            "fecha": n["fecha"], "bloque": n["bloque"],
            "tipo": n["tipo"], "descripcion": n["descripcion"],
            "creado_por": n["creado_por"],
        })
    ali_set       = {(a["empleado_id"], a["fecha"], a["bloque"]) for a in aliviadas}
    transporte_map = {s["empleado_id"]: s["saldo"] for s in saldos}

    grupos = {"TM": [], "TN": [], "CO": [], "ST": []}

    for emp in emp_rows:
        eid       = emp["id"]
        turno     = (emp["turno_nombre"] or "").lower()
        cortado   = emp["hipo"] == "cortado"
        if cortado:
            grupo = "CO"
        elif not emp["turno_nombre"]:
            if not emp["activo"]:
                continue
            grupo = "ST"
        elif "noche" in turno or "madrugada" in turno:
            grupo = "TN"
        else:
            grupo = "TM"

        f_ing = (emp["fecha_ingreso"] or "")[:10]
        f_egr = (emp["fecha_egreso"]  or "")[:10]

        celdas        = {}
        tots          = defaultdict(float)
        feriados_trab = 0.0

        for fecha in fechas:
            if cortado:
                b1 = _resolver(eid, fecha, 1, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                b2 = _resolver(eid, fecha, 2, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                celdas[fecha] = {"b1": b1, "b2": b2}
                for b in (b1, b2):
                    l = b["letra"]
                    if l and b["tipo"] == "normal" and l != "A!!!":
                        tots[l] += 0.5
                        if l in CONTABLES:
                            tots["dias"] += 0.5
                if fecha in feriados:
                    letras_feri = [b["letra"] for b in (b1, b2)
                                   if b.get("tipo") == "normal" and b.get("letra") in ("I", "T", "FT", "@", "NF")]
                    if any(l != "@" for l in letras_feri):
                        feriados_trab += len(letras_feri) * 0.5
            else:
                plan = plan_map.get((eid, fecha))
                dia_cortado = plan and plan.get("horario_tipo") == "cortado"
                if dia_cortado:
                    b1 = _resolver(eid, fecha, 1, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                    b2 = _resolver(eid, fecha, 2, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                    celdas[fecha] = {"b1": b1, "b2": b2, "es_cortado_dia": True}
                    for b in (b1, b2):
                        l = b["letra"]
                        if l and b["tipo"] == "normal" and l != "A!!!":
                            tots[l] += 0.5
                            if l in CONTABLES:
                                tots["dias"] += 0.5
                    if fecha in feriados:
                        letras_feri = [b["letra"] for b in (b1, b2)
                                       if b.get("tipo") == "normal" and b.get("letra") in ("I", "T", "FT", "@", "NF")]
                        if any(l != "@" for l in letras_feri):
                            feriados_trab += len(letras_feri) * 0.5
                else:
                    c = _resolver(eid, fecha, 0, f_ing, f_egr, nov_map, ali_set, res_map, plan_map)
                    celdas[fecha] = c
                    l = c["letra"]
                    if l and c["tipo"] == "normal" and l != "A!!!":
                        tots[l] += 1.0
                        if l in CONTABLES:
                            tots["dias"] += 1.0
                        if fecha in feriados and l in ("I", "T", "FT", "NF"):
                            feriados_trab += 1.0

        transporte = transporte_map.get(eid, 0)
        saldo_mes  = transporte + tots.get("FD", 0) - tots.get("FT", 0)
        control    = _calcular_control(fechas, f_egr, celdas, cortado, f0, f1,
                                       date.today().isoformat())

        grupos[grupo].append({
            "id":         eid,
            "cod":        emp["user_id"],
            "nombre":     emp["nombre"],
            "apellido":   emp["apellido"],
            "cargo":      emp["cargo"] or "",
            "fecha_ingreso": f_ing or None,
            "fecha_egreso": f_egr or None,
            "celdas":     celdas,
            "control":    control,
            "comentarios": comentarios_map.get(eid, []),
            "totales": {
                **{k: _fmt(v) for k, v in tots.items()},
                "transporte":         _fmt(transporte),
                "saldo_francos":      _fmt(saldo_mes),
                "feriados_trabajados": _fmt(feriados_trab),
                "FM":                 fm_count.get(eid, 0),
            },
        })

    return {
        "mes":  mes,
        "dias": _build_dias(fechas, feriados),
        "TM":   grupos["TM"],
        "TN":   grupos["TN"],
        "CO":   grupos["CO"],
        "ST":   grupos["ST"],
    }


# ── Endpoints ─────────────────────────────────────────────────────────────────
@router.get("/mensual")
def asistencia_mensual(mes: str = Query(None), _user=Depends(require_permiso("asistencia", "ver"))):
    if not mes:
        mes = date.today().strftime("%Y-%m")
    return _asistencia_datos(mes)


@router.get("/mensual/excel")
def asistencia_mensual_excel(mes: str = Query(None), _user=Depends(require_permiso("asistencia", "ver_todos"))):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    if not mes:
        mes = date.today().strftime("%Y-%m")

    datos = _asistencia_datos(mes)

    with db_session() as conn:
        row = conn.execute("SELECT valor FROM configuracion WHERE clave='nombre_empresa'").fetchone()
        nombre_empresa = row["valor"].strip() if row and row["valor"] else ""

    todos = sorted(
        [{"emp": e, "cortado": False} for e in datos["TM"]] +
        [{"emp": e, "cortado": False} for e in datos["TN"]] +
        [{"emp": e, "cortado": True}  for e in datos["CO"]],
        key=lambda x: f"{x['emp']['apellido']} {x['emp']['nombre']}".lower()
    )

    COLORES = {
        "I":    ("FFFFFF", "000000"), "T":    ("FF8C00", "000000"),
        "F":    ("27AE60", "FFFFFF"), "FT":   ("27AE60", "FFFFFF"),
        "FD":   ("27AE60", "FFFFFF"), "A":    ("E74C3C", "FFFFFF"),
        "A!!!": ("7B0000", "FFFFFF"), "E":    ("FABF8F", "000000"),
        "ILT":  ("C4D79B", "000000"), "V":    ("95B3D7", "000000"),
        "L":    ("92CDDC", "000000"), "LSG":  ("B1A0C7", "000000"),
        "S":    ("DA9694", "000000"), "@":    ("FFC7CE", "000000"),
        "NF":   ("F4D03F", "7D4900"), "CP":   ("9CCC65", "1B5E20"),
        "MT":   ("E67E22", "FFFFFF"), "X":    ("808080", "FFFFFF"),
        "O":    ("404040", "FFFFFF"), "!":    ("8E44AD", "FFFFFF"),
    }

    RES_COLS   = ["saldo_francos","T","ILT","LSG","L","@","E","V","S","NF","FM","A","feriados_trabajados","dias"]
    RES_LABELS = {
        "saldo_francos":"SALDO","T":"T","ILT":"ILT","LSG":"LSG","L":"L",
        "@":"@","E":"E","V":"V","S":"S","NF":"NF","FM":"FM","A":"A",
        "feriados_trabajados":"FERI.","dias":"DÍAS",
    }

    dias = datos["dias"]
    año_n, mes_n = int(mes[:4]), int(mes[5:7])
    meses_es = ["enero","febrero","marzo","abril","mayo","junio",
                "julio","agosto","septiembre","octubre","noviembre","diciembre"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{meses_es[mes_n-1].capitalize()} {año_n}"

    # Fila de título con nombre de empresa y período
    titulo = f"{nombre_empresa}  —  Planilla {meses_es[mes_n-1].capitalize()} {año_n}" if nombre_empresa else f"Planilla {meses_es[mes_n-1].capitalize()} {año_n}"
    ws.append([titulo])
    ws.append([])  # fila vacía de separación
    # Los encabezados y datos empiezan en fila 3
    HDR_ROW = 3

    hdr_fill  = PatternFill("solid", fgColor="1A252F")
    dom_fill  = PatternFill("solid", fgColor="4A235A")
    fer_fill  = PatternFill("solid", fgColor="1A5276")
    hdr_font  = Font(bold=True, color="FFFFFF", size=8)

    # Estilo título
    tc = ws.cell(1, 1)
    tc.value = titulo
    tc.font = Font(bold=True, size=11, color="1A252F")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    def hcell(col, text, fill=None, width=None, halign="center"):
        c = ws.cell(HDR_ROW, col, text)
        c.fill = fill or hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=True)
        if width:
            ws.column_dimensions[c.column_letter].width = width
        return c

    col = 1
    hcell(col, "COD",              width=6);  col += 1
    hcell(col, "APELLIDO Y NOMBRE",width=30, halign="left"); col += 1

    day_start = col
    for dia in dias:
        fill = fer_fill if dia["feriado"] else dom_fill if dia["domingo"] else hdr_fill
        hcell(col, f"{dia['num']}\n{dia['dow']}", fill=fill, width=4.5)
        col += 1

    ws.column_dimensions[ws.cell(HDR_ROW, col).column_letter].width = 1
    col += 1  # separador

    for rc in RES_COLS:
        hcell(col, RES_LABELS[rc], width=5.5); col += 1

    ctrl_col = col
    hcell(col, "CONTROL", width=20, halign="left")

    ws.freeze_panes = f"C{HDR_ROW + 1}"
    ws.row_dimensions[HDR_ROW].height = 30

    def fmt_dias(arr):
        return ", ".join(str(x) for x in arr[:6]) + ("…" if len(arr) > 6 else "")

    for row_i, item in enumerate(todos, start=HDR_ROW + 1):
        emp = item["emp"]
        es_co = item["cortado"]

        c = ws.cell(row_i, 1, emp.get("cod") or "")
        c.alignment = Alignment(horizontal="center"); c.font = Font(size=8)

        c = ws.cell(row_i, 2, f"{emp['apellido']}, {emp['nombre']}")
        c.alignment = Alignment(horizontal="left"); c.font = Font(size=8)

        col = day_start
        for dia in dias:
            fecha  = dia["fecha"]
            celda  = emp["celdas"].get(fecha, {})
            if es_co or celda.get("es_cortado_dia"):
                b1 = celda.get("b1") or {"letra": None, "tipo": "sin_plan"}
                b2 = celda.get("b2") or {"letra": None, "tipo": "sin_plan"}
                l1, l2 = b1.get("letra") or "", b2.get("letra") or ""
                text     = f"{l1}/{l2}" if (l1 or l2) else ""
                bg_letra = l1 or l2
                tipo     = b1.get("tipo", "sin_plan")
            else:
                l        = celda.get("letra") or ""
                text     = l
                bg_letra = l
                tipo     = celda.get("tipo", "sin_plan")

            c = ws.cell(row_i, col, text)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if tipo == "antes_ingreso":
                c.fill = PatternFill("solid", fgColor="000000")
                c.font = Font(size=7, bold=True, color="000000")
            elif tipo == "liquidacion":
                c.fill = PatternFill("solid", fgColor="5D6D7E")
                c.font = Font(size=7, bold=True, color="FFFFFF")
            elif bg_letra in COLORES:
                bg, fg = COLORES[bg_letra]
                c.fill = PatternFill("solid", fgColor=bg)
                c.font = Font(size=7, bold=True, color=fg)
            else:
                c.font = Font(size=7)
            col += 1

        col += 1  # separador

        for rc in RES_COLS:
            v = (emp.get("totales") or {}).get(rc, 0) or 0
            c = ws.cell(row_i, col, v if v != 0 else "")
            c.alignment = Alignment(horizontal="center"); c.font = Font(size=8)
            col += 1

        ctrl = emp.get("control") or {}
        estado = ctrl.get("estado", "")
        dd, df, da = ctrl.get("dias_duda",[]), ctrl.get("dias_faltantes",[]), ctrl.get("dias_ausentes_sc",[])
        sc = f" · A!!! ({fmt_dias(da)})" if da else ""
        ctrl_txt = {
            "completado":  "COMPLETADO",
            "liquidacion": "LIQ. FINAL",
        }.get(estado, "")
        if not ctrl_txt:
            if estado in ("ausentes_sc","en_curso") and da:
                ctrl_txt = f"A!!! ({fmt_dias(da)})"
            elif estado == "revisar":
                ctrl_txt = f"Rev: {fmt_dias(dd)}{sc}"
            elif estado == "faltan":
                ctrl_txt = f"Falt: {fmt_dias(df)}{sc}"
            elif estado == "mixto":
                ctrl_txt = f"Rev: {fmt_dias(dd)} · Falt: {fmt_dias(df)}{sc}"

        CTRL_COLORES = {
            "completado":  ("D4EDDA", "155724"),
            "liquidacion": ("E74C3C", "FFFFFF"),
            "revisar":     ("FFF3CD", "7D5A00"),
            "faltan":      ("F8D7DA", "721C24"),
            "mixto":       ("F8D7DA", "721C24"),
            "ausentes_sc": ("7B0000", "FFFFFF"),
        }
        c = ws.cell(row_i, ctrl_col, ctrl_txt)
        c.alignment = Alignment(horizontal="left"); c.font = Font(size=7)
        if estado in CTRL_COLORES:
            bg, fg = CTRL_COLORES[estado]
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(size=7, color=fg)
        elif estado in ("en_curso",) and da:
            c.fill = PatternFill("solid", fgColor="7B0000")
            c.font = Font(size=7, color="FFFFFF")
        ws.row_dimensions[row_i].height = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    prefix = f"{nombre_empresa.replace(' ', '_')}_" if nombre_empresa else ""
    filename = f"{prefix}planilla_{meses_es[mes_n-1]}_{año_n}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Novedades CRUD ────────────────────────────────────────────────────────────
class NovedadIn(BaseModel):
    empleado_id: int
    fecha:       str
    bloque:      int = 0
    tipo:        str
    descripcion: Optional[str] = None


@router.post("/novedades", status_code=201)
def upsert_novedad(data: NovedadIn, user=Depends(require_permiso("asistencia", "corregir"))):
    if data.tipo not in LETRAS_VALIDAS:
        raise HTTPException(400, f"Tipo inválido. Válidos: {sorted(LETRAS_VALIDAS)}")
    if data.tipo == "@" and data.bloque not in (1, 2):
        raise HTTPException(400, "La aliviada (@) requiere bloque 1 o 2")
    if data.tipo == "CP" and not data.descripcion:
        raise HTTPException(400, "La compensación de presencia requiere una descripción")
    if data.tipo == "V":
        from api.vacaciones import _calcular_dias_formula, _get_arrastre
        anio_periodo = int(data.fecha[:4]) - 1
        with db_session() as conn:
            emp = conn.execute("SELECT fecha_ingreso FROM empleados WHERE id=?", (data.empleado_id,)).fetchone()
            saldos = {(r["empleado_id"], r["anio"]): dict(r)
                      for r in conn.execute("SELECT * FROM vacaciones_saldo_inicial").fetchall()}
            dias_v_ya = conn.execute("""
                SELECT SUM(CASE WHEN bloque=0 THEN 1.0 ELSE 0.5 END)
                FROM novedades
                WHERE tipo='V' AND empleado_id=? AND strftime('%Y',fecha)=?
                  AND NOT (fecha=? AND bloque=?)
            """, (data.empleado_id, str(int(data.fecha[:4])),
                  data.fecha, data.bloque)).fetchone()[0] or 0.0
        fi = emp["fecha_ingreso"] if emp else None
        _, dias_formula = _calcular_dias_formula(fi, anio_periodo)
        saldo = saldos.get((data.empleado_id, anio_periodo))
        nueva = 0.5 if data.bloque > 0 else 1.0
        if saldo:
            dias_restan = saldo["dias_correspondian"] - saldo["dias_tomados"] - dias_v_ya
        else:
            arrastre = _get_arrastre(data.empleado_id, anio_periodo, fi, saldos,
                                     {(data.empleado_id, int(data.fecha[:4])): {"01": dias_v_ya} if dias_v_ya else {}})
            dias_restan = dias_formula + arrastre - dias_v_ya
        dias_restan = round(dias_restan, 1)
        if dias_restan < nueva:
            raise HTTPException(400, f"Sin días de vacaciones disponibles. Quedan {dias_restan:.1f} día(s).")
    creado_por = user.get("email") or user.get("sub")
    from api.periodos_cerrados import check_periodo_abierto
    with db_session() as conn:
        check_periodo_abierto(conn, data.fecha)
        conn.execute(
            """INSERT INTO novedades (empleado_id, fecha, bloque, tipo, descripcion, creado_por)
               VALUES (?,?,?,?,?,?)
               ON CONFLICT(empleado_id, fecha, bloque)
               DO UPDATE SET tipo=excluded.tipo, descripcion=excluded.descripcion,
                             creado_por=excluded.creado_por""",
            (data.empleado_id, data.fecha, data.bloque, data.tipo, data.descripcion, creado_por),
        )
        row = conn.execute(
            "SELECT * FROM novedades WHERE empleado_id=? AND fecha=? AND bloque=?",
            (data.empleado_id, data.fecha, data.bloque),
        ).fetchone()
    return dict(row)


class NovedadRangoIn(BaseModel):
    empleado_id: int
    fecha_desde: str
    fecha_hasta: str
    bloque:      int = 0


# Ruta estática antes que la parametrizada para evitar que {nov_id} la capture
@router.post("/novedades/borrar-rango", status_code=200)
def eliminar_novedades_rango(data: NovedadRangoIn, _user=Depends(require_permiso("asistencia", "corregir"))):
    if data.fecha_hasta < data.fecha_desde:
        raise HTTPException(400, "fecha_hasta debe ser >= fecha_desde")
    from api.periodos_cerrados import check_rango_abierto
    with db_session() as conn:
        check_rango_abierto(conn, data.fecha_desde, data.fecha_hasta)
        tipos_afectados = {
            r["tipo"] for r in conn.execute(
                "SELECT DISTINCT tipo FROM novedades WHERE empleado_id=? AND fecha>=? AND fecha<=? AND bloque=?",
                (data.empleado_id, data.fecha_desde, data.fecha_hasta, data.bloque),
            ).fetchall()
        }
        conn.execute(
            """DELETE FROM novedades
               WHERE empleado_id=? AND fecha>=? AND fecha<=? AND bloque=?""",
            (data.empleado_id, data.fecha_desde, data.fecha_hasta, data.bloque),
        )
    if tipos_afectados & {"CP", "NF"}:
        try:
            from sync.evaluador import evaluar_rango
            evaluar_rango(data.fecha_desde, data.fecha_hasta, respetar_correcciones=False)
        except Exception:
            pass
    return {"ok": True}


@router.delete("/novedades/{nov_id}")
def eliminar_novedad(nov_id: int, _user=Depends(require_permiso("asistencia", "corregir"))):
    from api.periodos_cerrados import check_periodo_abierto
    with db_session() as conn:
        row = conn.execute("SELECT id, fecha, tipo, empleado_id FROM novedades WHERE id=?", (nov_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Novedad no encontrada")
        check_periodo_abierto(conn, row["fecha"])
        conn.execute("DELETE FROM novedades WHERE id=?", (nov_id,))
        tipo, fecha, empleado_id = row["tipo"], row["fecha"], row["empleado_id"]
    if tipo in ("CP", "NF"):
        try:
            from sync.evaluador import evaluar_fecha
            evaluar_fecha(fecha, respetar_correcciones=False, solo_empleado_id=empleado_id)
        except Exception:
            pass
    return {"ok": True}


# ── Saldo francos (TRANSPORTE carry-forward) ──────────────────────────────────
@router.put("/saldo_francos/{empleado_id}/{mes}")
def set_saldo_francos(empleado_id: int, mes: str, body: dict, _user=Depends(require_permiso("asistencia", "corregir"))):
    saldo = float(body.get("saldo", 0))
    from api.periodos_cerrados import check_periodo_abierto
    with db_session() as conn:
        check_periodo_abierto(conn, f"{mes}-01")
        conn.execute(
            """INSERT INTO saldo_francos (empleado_id, mes, saldo) VALUES (?,?,?)
               ON CONFLICT(empleado_id, mes) DO UPDATE SET saldo=excluded.saldo""",
            (empleado_id, mes, saldo),
        )
    return {"ok": True}


# ── Orden de planilla ─────────────────────────────────────────────────────────
class OrdenItem(BaseModel):
    empleado_id: Optional[int] = None
    etiqueta:    Optional[str] = None


@router.get("/orden/{grupo}")
def get_orden(grupo: str, mes: str | None = None, _user=Depends(require_permiso("asistencia", "ver"))):
    if grupo not in ("TM", "TN", "CO"):
        raise HTTPException(400, "Grupo inválido")
    with db_session() as conn:
        rows = conn.execute(
            "SELECT id, posicion, empleado_id, etiqueta FROM planilla_orden "
            "WHERE grupo=? ORDER BY posicion",
            (grupo,),
        ).fetchall()

        if mes:
            try:
                from datetime import date
                y, m = map(int, mes.split("-"))
                import calendar
                f0 = f"{y}-{m:02d}-01"
                f1 = f"{y}-{m:02d}-{calendar.monthrange(y, m)[1]:02d}"
            except Exception:
                return [dict(r) for r in rows]

            # Determinar grupo real de cada empleado para ese mes
            emp_grupo = conn.execute("""
                WITH ult AS (
                    SELECT p.empleado_id,
                           h.tipo, t.nombre AS turno_nombre,
                           ROW_NUMBER() OVER (
                               PARTITION BY p.empleado_id ORDER BY p.fecha DESC
                           ) AS rn
                    FROM planificacion p
                    JOIN horarios h ON h.id = p.horario_id
                    LEFT JOIN turnos t ON t.id = h.turno_id
                    WHERE p.fecha >= ? AND p.fecha <= ? AND p.horario_id IS NOT NULL
                )
                SELECT empleado_id,
                    CASE
                        WHEN tipo = 'cortado' THEN 'CO'
                        WHEN lower(COALESCE(turno_nombre,'')) LIKE '%noche%'
                          OR lower(COALESCE(turno_nombre,'')) LIKE '%madrugada%' THEN 'TN'
                        ELSE 'TM'
                    END AS grupo
                FROM ult WHERE rn = 1
            """, (f0, f1)).fetchall()
            ids_del_grupo = {r["empleado_id"] for r in emp_grupo if r["grupo"] == grupo}

            rows = [r for r in rows
                    if r["empleado_id"] is None or r["empleado_id"] in ids_del_grupo]

    return [dict(r) for r in rows]


@router.post("/orden/{grupo}", status_code=200)
def set_orden(grupo: str, items: list[OrdenItem], _user=Depends(require_permiso("asistencia", "editar"))):
    if grupo not in ("TM", "TN", "CO"):
        raise HTTPException(400, "Grupo inválido")
    with db_session() as conn:
        conn.execute("DELETE FROM planilla_orden WHERE grupo=?", (grupo,))
        for pos, item in enumerate(items):
            conn.execute(
                "INSERT INTO planilla_orden (grupo, posicion, empleado_id, etiqueta) VALUES (?,?,?,?)",
                (grupo, pos, item.empleado_id, item.etiqueta),
            )
    return {"ok": True}
